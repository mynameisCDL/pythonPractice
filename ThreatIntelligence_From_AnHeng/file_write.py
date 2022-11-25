from openpyxl import load_workbook
import shutil
import os
import time
from colorama import Fore


def file_creat(file_name):
    file_name = file_name.replace(':', '')
    if not os.path.exists('result-{}/'.format(time.strftime("%Y%m%d", time.localtime())) + file_name):
        os.makedirs('result-{}/'.format(time.strftime("%Y%m%d", time.localtime())) + file_name)
        os.makedirs('result-{}/'.format(time.strftime("%Y%m%d", time.localtime())) + file_name + '/' + '附件')
    result_file = 'result-{}/'.format(time.strftime("%Y%m%d", time.localtime())) + file_name + '/' + file_name + '.xlsx'
    shutil.copy("模板/模板.xlsx", result_file)

    return result_file


def basic_info_write(data, result_file):
    wb = load_workbook(result_file)
    ws = wb['ReportDetail']

    ws['B1'] = data['title']
    ws['B2'] = data['text']
    ws['B4'] = data['url']
    ws['B6'] = data['create_time']

    wb.save(result_file)
    print(Fore.GREEN + '[{}] [INFO] 基础信息 写入文件完成！'.format(time.strftime("%H:%M:%S", time.localtime())))


def ioc_info_write(ioc_list, result_file):
    wb = load_workbook(result_file)
    ws = wb['IOC']

    row = 2
    for ioc in ioc_list:
        # 给关键字重命名，因为填入表里的ioc类型名称有规定
        if ioc.get("keyword") == "domain":
            keyword = "Domain"
        elif ioc.get("keyword") == "ip":
            keyword = "IPV4 address"
        else:
            keyword = "SHA256"

        info = ioc.get("info")
        tag = ioc.get("tag")
        ws['A' + str(row)] = keyword
        ws['B' + str(row)] = info
        ws['C' + str(row)] = tag
        row = row + 1

    wb.save(result_file)
    print(Fore.GREEN + '[{}] [INFO] IOC信息 写入文件完成！'.format(time.strftime("%H:%M:%S", time.localtime())))
